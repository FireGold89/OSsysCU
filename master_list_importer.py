"""
master_list_importer.py — 公司 Master List（Quotation & Contract number）匯入

Phase 1：QS 主檔欄位；Phase 2：財務 Admin 欄（糧期／分判付款／支票子表）。
唯一鍵：完整報價編號，例 MS/Q001/26/jy

獨立執行:
  python master_list_importer.py preview <file.xlsx>
  python master_list_importer.py sync <file.xlsx>
  python master_list_importer.py sync-all   # Ref 資料夾（預設）2017→2026
"""
import json
import os
import re
import sys
from datetime import datetime

import openpyxl

import database as db
from master_ref import enrich_person_fields, extract_person_code_from_quotation_no, normalize_person_code

# Phase 1 比對欄位（財務欄不在此列）
SYNC_FIELDS = (
    'quote_date', 'doc_type', 'awarded', 'site_name', 'trade_category',
    'description', 'person_code', 'person_in_charge', 'client_name', 'quoted_amount',
    'margin_pct', 'awarded_amount', 'contract_days', 'start_date',
    'completion_date', 'subcon_type', 'subcon_company', 'subcon_amount',
    'profit_amount', 'profit_pct',
)

SKIP_SHEET_KEYWORDS = ('分類', '投標額度', '500 萬', '500萬')
MASTER_DATA_SHEETS = frozenset(('報價', '標書', '合約', '中標項目'))


def safe_str(val):
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ('none', 'nan'):
        return None
    if s.upper() in ('NA', 'N/A', '-', '—'):
        return None
    return s


def safe_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s or s.upper() in ('NA', 'N/A', '-', '—', 'NONE', 'NAN'):
        return None
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y'):
        try:
            return datetime.strptime(s[:10], fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    if re.match(r'\d{4}-\d{2}-\d{2}', s):
        return s[:10]
    return None


def safe_float(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s.upper() in ('NA', 'N/A', '-', '—', '無意投報', '單價項目'):
        return None
    m = re.search(r'[\d,]+(?:\.\d+)?', s.replace(',', ''))
    if m:
        try:
            return float(m.group().replace(',', ''))
        except ValueError:
            pass
    try:
        v = float(s.replace(',', '').replace('$', ''))
        return v
    except ValueError:
        return None


def safe_pct(val):
    v = safe_float(val)
    if v is None:
        return None
    if abs(v) <= 1.5:
        return round(v * 100, 4)
    return round(v, 4)


def build_quotation_no(p1, p2, p3, p4=None):
    """合併 B–E 欄 → MS/Q001/26/jy 或 MS/C001/17"""
    p4_use = p4 if _is_person_suffix(p4) else None
    parts = [safe_str(x) for x in (p1, p2, p3, p4_use)]
    parts = [p for p in parts if p]
    if len(parts) < 2:
        return None
    q = parts[0]
    for p in parts[1:]:
        if p.startswith('/'):
            q = q.rstrip('/') + p
        elif q.endswith('/'):
            q += p
        else:
            q += '/' + p
    q = re.sub(r'/+', '/', q)
    if not re.search(r'[QTC]\d{3,}', q, re.I):
        return None
    return q


def is_master_list_file(name):
    """略過 Excel 暫存檔 ~$"""
    if name.startswith('~$') or name.startswith('~'):
        return False
    nl = name.lower()
    return nl.endswith('.xlsx') and ('quotation' in nl or 'contract number' in nl)


def normalize_master_source_file(name):
    """正規 Master List 檔名 → YYYY Quotation & Contract number.xlsx"""
    base = os.path.basename(name or '')
    if not base:
        return base
    m = re.search(r'(20\d{2})', base)
    if m and is_master_list_file(base):
        return f"{m.group(1)} Quotation & Contract number.xlsx"
    return base


def is_master_data_sheet(name):
    return name in MASTER_DATA_SHEETS


def _ordered_master_sheets(sheetnames):
    """先讀報價／合約，最後合併「中標項目」（避免被報價列覆蓋中標標記）"""
    priority = {'報價': 0, '標書': 1, '合約': 2, '中標項目': 9}
    return sorted(
        [s for s in sheetnames if is_master_data_sheet(s)],
        key=lambda s: priority.get(s, 5),
    )


def _header_cell(row_vals, idx):
    if idx >= len(row_vals):
        return ''
    return str(row_vals[idx] or '').replace('\n', ' ').strip()


def _is_header_row(row_vals):
    head = ''.join(str(c or '') for c in row_vals[:8]).lower()
    if '報價編號' in head or 'quotation' in head:
        return True
    c0 = str(row_vals[0] or '').strip().lower()
    return c0 in ('日期', 'date')


def _is_person_suffix(val):
    """E 欄是否為負責人縮寫（合約 sheet 的 E 欄常為屋苑名稱）"""
    s = safe_str(val)
    if not s:
        return False
    if s.startswith('/'):
        return True
    return bool(re.match(r'^[a-z]{2,4}$', s, re.I))


def detect_quote_layout(header_vals):
    """
    依表頭辨識報價／標書版型（以 2025–2026 為 modern 藍本）：
    - modern (2022+)：F=中標、G=報價/標書
    - transitional_2021：F=報價/標書、G=屋苑/地點（無中標欄）
    - transitional_site (2019–2020)：F=屋苑/地點、G=工作範疇
    - legacy_2018 (2018)：F=Site Name、G=工作範疇、J=業主、N=外判
    - legacy_2017 (2017)：F=Site Name、G=Description、I=業主
    """
    f = _header_cell(header_vals, 5).lower()
    g = _header_cell(header_vals, 6).lower()
    if '中標' in f and ('報價' in g or '標書' in g):
        return 'modern'
    if ('報價' in f or '標書' in f) and '屋苑' in g:
        return 'transitional_2021'
    if '屋苑' in f and '工作範疇' in g:
        return 'transitional_site'
    if 'site name' in f:
        if '工作範疇' in g:
            return 'legacy_2018'
        return 'legacy_2017'
    return 'legacy_2017'


def detect_contract_layout(header_vals):
    """合約 sheet：2018+ F=總價目；2017 F=Description"""
    f = _header_cell(header_vals, 5).lower()
    if '總價' in f:
        return 'contract_site_amount'
    return 'contract_legacy'


def map_header_columns(header_vals):
    """
    由表頭找出分判欄索引（外判公司→分判商、外判金額/分判金額→分判金額）。
    略過 Admin 填寫欄，優先使用 QS 欄（2022+ 的 X/Y）。
    """
    m = {}
    for i, raw in enumerate(header_vals):
        h = str(raw or '').replace('\n', ' ').strip().lower()
        if not h or 'admin' in h or '由admin' in h:
            continue
        if '外判與否' in h and 'subcon_type' not in m:
            m['subcon_type'] = i
        elif ('主要分判' in h or '外判公司' in h) and 'subcon_company' not in m:
            m['subcon_company'] = i
        elif ('分判金額' in h or '外判金額' in h) and 'subcon_amount' not in m:
            m['subcon_amount'] = i
    return m


def _pick_col(row_vals, col_map, key, default_idx, parser=safe_str):
    idx = col_map.get(key, default_idx) if col_map else default_idx
    if idx is None or idx >= len(row_vals):
        return None
    return parser(row_vals[idx])


def apply_subcon_fields(rec, row_vals, col_map, defaults=None):
    """defaults: {subcon_type: 22, subcon_company: 23, subcon_amount: 24}"""
    d = defaults or {}
    rec['subcon_type'] = _pick_col(row_vals, col_map, 'subcon_type', d.get('subcon_type'))
    rec['subcon_company'] = _pick_col(row_vals, col_map, 'subcon_company', d.get('subcon_company'))
    rec['subcon_amount'] = _pick_col(
        row_vals, col_map, 'subcon_amount', d.get('subcon_amount'), safe_float,
    )
    return rec


def infer_source_year(filepath, year_suffix=None):
    m = re.search(r'(20\d{2})', os.path.basename(filepath))
    if m:
        return int(m.group(1))
    if year_suffix and re.match(r'/(\d{2})$', year_suffix):
        yy = int(year_suffix[1:])
        return 2000 + yy if yy < 80 else 1900 + yy
    return None


def _base_record(quotation_no, source_file, source_sheet, row_vals):
    year_suffix = safe_str(row_vals[3]) if len(row_vals) > 3 else None
    rec = {
        'quotation_no': quotation_no,
        'source_year': infer_source_year(source_file, year_suffix),
        'quote_date': safe_date(row_vals[0]),
        'source_file': normalize_master_source_file(source_file),
        'source_sheet': source_sheet,
        'person_code': normalize_person_code(row_vals[4]) if len(row_vals) > 4 else None,
    }
    return enrich_person_fields(rec)


def parse_modern_row(row_vals, source_file, source_sheet, col_map=None, finance_map=None):
    """2022–2026 標準 Master List（F=中標、G=報價/標書）"""
    if len(row_vals) < 8:
        return None
    quotation_no = build_quotation_no(row_vals[1], row_vals[2], row_vals[3], row_vals[4])
    if not quotation_no:
        return None

    doc_type = safe_str(row_vals[6]) or ('標書' if source_sheet == '標書' else '報價')
    awarded_raw = safe_str(row_vals[5])
    awarded = '中' if awarded_raw == '中' else None

    checklist = {
        'bid_signoff': safe_str(row_vals[16]) if len(row_vals) > 16 else None,
        'partner_form': safe_str(row_vals[17]) if len(row_vals) > 17 else None,
        'contract_signoff': safe_str(row_vals[18]) if len(row_vals) > 18 else None,
    }
    checklist = {k: v for k, v in checklist.items() if v}

    trade = safe_str(row_vals[9]) if len(row_vals) > 9 else None
    if not trade:
        trade = safe_str(row_vals[8]) if len(row_vals) > 8 else None

    contract_days = safe_float(row_vals[19]) if len(row_vals) > 19 else None
    if contract_days is not None:
        contract_days = int(contract_days)

    rec = _base_record(quotation_no, source_file, source_sheet, row_vals)
    rec.update({
        'doc_type': doc_type,
        'awarded': awarded,
        'site_name': safe_str(row_vals[7]) if len(row_vals) > 7 else None,
        'trade_category': trade,
        'description': safe_str(row_vals[10]) if len(row_vals) > 10 else None,
        'person_in_charge': safe_str(row_vals[11]) if len(row_vals) > 11 else rec.get('person_in_charge'),
        'client_name': safe_str(row_vals[12]) if len(row_vals) > 12 else None,
        'quoted_amount': safe_float(row_vals[13]) if len(row_vals) > 13 else None,
        'margin_pct': safe_pct(row_vals[14]) if len(row_vals) > 14 else None,
        'awarded_amount': safe_float(row_vals[15]) if len(row_vals) > 15 else None,
        'contract_days': contract_days,
        'start_date': safe_date(row_vals[20]) if len(row_vals) > 20 else None,
        'completion_date': safe_date(row_vals[21]) if len(row_vals) > 21 else None,
        'subcon_type': None,
        'subcon_company': None,
        'subcon_amount': None,
        'profit_amount': None,
        'profit_pct': None,
        'checklist_json': json.dumps(checklist, ensure_ascii=False) if checklist else None,
    })
    apply_subcon_fields(rec, row_vals, col_map, {'subcon_type': 22})
    from master_finance import apply_qs_subcon_to_record, extract_finance_detail
    apply_qs_subcon_to_record(rec, row_vals, finance_map)
    finance = extract_finance_detail(row_vals, finance_map, 'modern')
    if any(finance.get(k) for k in ('summary', 'qs_subcon_lines', 'client_invoices', 'subcon_payments', 'cheques')):
        rec['_finance'] = finance
    return enrich_person_fields(rec)


def parse_transitional_site_row(row_vals, source_file, source_sheet, col_map=None):
    """2019–2020：F=屋苑/地點、G=工作範疇（無中標／報價類型欄）"""
    if len(row_vals) < 8:
        return None
    quotation_no = build_quotation_no(row_vals[1], row_vals[2], row_vals[3], row_vals[4])
    if not quotation_no:
        return None

    doc_type = '標書' if source_sheet == '標書' else '報價'
    rec = _base_record(quotation_no, source_file, source_sheet, row_vals)
    rec.update({
        'doc_type': doc_type,
        'awarded': None,
        'site_name': safe_str(row_vals[5]),
        'trade_category': safe_str(row_vals[6]) if len(row_vals) > 6 else None,
        'description': safe_str(row_vals[7]) if len(row_vals) > 7 else None,
        'person_in_charge': safe_str(row_vals[8]) if len(row_vals) > 8 else rec.get('person_in_charge'),
        'client_name': safe_str(row_vals[9]) if len(row_vals) > 9 else None,
        'quoted_amount': safe_float(row_vals[10]) if len(row_vals) > 10 else None,
        'margin_pct': None,
        'awarded_amount': None,
        'contract_days': None,
        'start_date': safe_date(row_vals[11]) if len(row_vals) > 11 else None,
        'completion_date': safe_date(row_vals[12]) if len(row_vals) > 12 else None,
        'subcon_type': None,
        'subcon_company': None,
        'subcon_amount': None,
        'profit_amount': None,
        'profit_pct': None,
        'checklist_json': None,
    })
    apply_subcon_fields(rec, row_vals, col_map, {'subcon_type': 13})
    return enrich_person_fields(rec)


def parse_transitional_2021_row(row_vals, source_file, source_sheet, col_map=None):
    """2021：F=報價/標書、G=屋苑/地點（無中標欄）"""
    if len(row_vals) < 8:
        return None
    quotation_no = build_quotation_no(row_vals[1], row_vals[2], row_vals[3], row_vals[4])
    if not quotation_no:
        return None

    doc_type = safe_str(row_vals[5]) or ('標書' if source_sheet == '標書' else '報價')
    rec = _base_record(quotation_no, source_file, source_sheet, row_vals)
    rec.update({
        'doc_type': doc_type,
        'awarded': None,
        'site_name': safe_str(row_vals[6]) if len(row_vals) > 6 else None,
        'trade_category': safe_str(row_vals[7]) if len(row_vals) > 7 else None,
        'description': safe_str(row_vals[8]) if len(row_vals) > 8 else None,
        'person_in_charge': safe_str(row_vals[9]) if len(row_vals) > 9 else rec.get('person_in_charge'),
        'client_name': safe_str(row_vals[10]) if len(row_vals) > 10 else None,
        'quoted_amount': safe_float(row_vals[11]) if len(row_vals) > 11 else None,
        'margin_pct': None,
        'awarded_amount': None,
        'contract_days': None,
        'start_date': safe_date(row_vals[14]) if len(row_vals) > 14 else None,
        'completion_date': safe_date(row_vals[15]) if len(row_vals) > 15 else None,
        'subcon_type': None,
        'subcon_company': None,
        'subcon_amount': None,
        'profit_amount': None,
        'profit_pct': None,
        'checklist_json': None,
    })
    apply_subcon_fields(rec, row_vals, col_map, {'subcon_type': 16})
    return enrich_person_fields(rec)


def parse_legacy_2017_row(row_vals, source_file, source_sheet):
    """2017：I=業主、M=外判與否、S=外判公司(分判商)、T=外判金額(分判金額)"""
    if len(row_vals) < 8:
        return None
    quotation_no = build_quotation_no(row_vals[1], row_vals[2], row_vals[3], row_vals[4])
    if not quotation_no:
        return None

    rec = _base_record(quotation_no, source_file, source_sheet, row_vals)
    rec.update({
        'doc_type': '報價',
        'awarded': None,
        'site_name': safe_str(row_vals[5]),
        'trade_category': None,
        'description': safe_str(row_vals[6]),
        'person_in_charge': safe_str(row_vals[7]) or rec.get('person_in_charge'),
        'client_name': safe_str(row_vals[8]) if len(row_vals) > 8 else None,
        'quoted_amount': safe_float(row_vals[9]) if len(row_vals) > 9 else None,
        'margin_pct': None,
        'awarded_amount': None,
        'contract_days': None,
        'start_date': safe_date(row_vals[10]) if len(row_vals) > 10 else None,
        'completion_date': safe_date(row_vals[11]) if len(row_vals) > 11 else None,
        'subcon_type': safe_str(row_vals[12]) if len(row_vals) > 12 else None,
        'subcon_company': safe_str(row_vals[18]) if len(row_vals) > 18 else None,
        'subcon_amount': safe_float(row_vals[19]) if len(row_vals) > 19 else None,
        'profit_amount': None,
        'profit_pct': None,
        'checklist_json': None,
    })
    from master_finance import apply_master_profit_fields
    apply_master_profit_fields(rec)
    return enrich_person_fields(rec)


def parse_legacy_2018_row(row_vals, source_file, source_sheet):
    """2018：J=業主、N=外判與否、T=外判公司(分判商)、U=外判金額(分判金額)"""
    if len(row_vals) < 9:
        return None
    quotation_no = build_quotation_no(row_vals[1], row_vals[2], row_vals[3], row_vals[4])
    if not quotation_no:
        return None

    rec = _base_record(quotation_no, source_file, source_sheet, row_vals)
    rec.update({
        'doc_type': '報價',
        'awarded': None,
        'site_name': safe_str(row_vals[5]),
        'trade_category': safe_str(row_vals[6]) if len(row_vals) > 6 else None,
        'description': safe_str(row_vals[7]) if len(row_vals) > 7 else None,
        'person_in_charge': safe_str(row_vals[8]) if len(row_vals) > 8 else rec.get('person_in_charge'),
        'client_name': safe_str(row_vals[9]) if len(row_vals) > 9 else None,
        'quoted_amount': safe_float(row_vals[10]) if len(row_vals) > 10 else None,
        'margin_pct': None,
        'awarded_amount': None,
        'contract_days': None,
        'start_date': safe_date(row_vals[11]) if len(row_vals) > 11 else None,
        'completion_date': safe_date(row_vals[12]) if len(row_vals) > 12 else None,
        'subcon_type': safe_str(row_vals[13]) if len(row_vals) > 13 else None,
        'subcon_company': safe_str(row_vals[19]) if len(row_vals) > 19 else None,
        'subcon_amount': safe_float(row_vals[20]) if len(row_vals) > 20 else None,
        'profit_amount': None,
        'profit_pct': None,
        'checklist_json': None,
    })
    from master_finance import apply_master_profit_fields
    apply_master_profit_fields(rec)
    return enrich_person_fields(rec)


def parse_legacy_quote_row(row_vals, source_file, source_sheet):
    """舊版英文欄位後備（同 2017）"""
    return parse_legacy_2017_row(row_vals, source_file, source_sheet)


def parse_contract_site_amount_row(row_vals, source_file, source_sheet):
    """2018–2019 合約：E=Site、F=總價目、G=Description、H=Issue By"""
    if len(row_vals) < 6:
        return None
    quotation_no = build_quotation_no(row_vals[1], row_vals[2], row_vals[3], row_vals[4])
    if not quotation_no:
        return None

    site_name = safe_str(row_vals[4]) if len(row_vals) > 4 and not _is_person_suffix(row_vals[4]) else None

    rec = _base_record(quotation_no, source_file, source_sheet, row_vals)
    rec.update({
        'doc_type': '合約',
        'awarded': '中',
        'site_name': site_name,
        'trade_category': None,
        'description': safe_str(row_vals[6]) if len(row_vals) > 6 else None,
        'person_in_charge': safe_str(row_vals[7]) if len(row_vals) > 7 else None,
        'client_name': safe_str(row_vals[8]) if len(row_vals) > 8 else None,
        'quoted_amount': None,
        'margin_pct': None,
        'awarded_amount': safe_float(row_vals[5]) if len(row_vals) > 5 else None,
        'contract_days': None,
        'start_date': None,
        'completion_date': None,
        'subcon_type': None,
        'subcon_company': safe_str(row_vals[8]) if len(row_vals) > 8 else None,
        'subcon_amount': None,
        'profit_amount': None,
        'profit_pct': None,
        'checklist_json': None,
    })
    return enrich_person_fields(rec)


def parse_legacy_contract_row(row_vals, source_file, source_sheet):
    """2017 合約 sheet（E=Project Name、F=Description）"""
    if len(row_vals) < 6:
        return None
    quotation_no = build_quotation_no(row_vals[1], row_vals[2], row_vals[3], row_vals[4])
    if not quotation_no:
        return None

    site_name = safe_str(row_vals[4]) if len(row_vals) > 4 else None
    if not site_name:
        site_name = safe_str(row_vals[5])
    description = safe_str(row_vals[5]) if len(row_vals) > 5 else None
    if len(row_vals) > 6 and safe_str(row_vals[6]):
        description = safe_str(row_vals[6])
    person = safe_str(row_vals[7]) if len(row_vals) > 7 else safe_str(row_vals[6])

    rec = _base_record(quotation_no, source_file, source_sheet, row_vals)
    rec.update({
        'doc_type': '合約',
        'awarded': '中',
        'site_name': site_name,
        'trade_category': None,
        'description': description,
        'person_in_charge': person,
        'client_name': None,
        'quoted_amount': None,
        'margin_pct': None,
        'awarded_amount': None,
        'contract_days': None,
        'start_date': None,
        'completion_date': None,
        'subcon_type': None,
        'subcon_company': None,
        'subcon_amount': None,
        'profit_amount': None,
        'profit_pct': None,
        'checklist_json': None,
    })
    return enrich_person_fields(rec)


def parse_awarded_sheet_row(row_vals, source_file, source_sheet):
    """
    「中標項目」工作表（2019–2021）：A–D = 報價編號，無日期欄。
    此表為該年中標清單；2022+ 改在報價 sheet F 欄標「中」。
    """
    if len(row_vals) < 5:
        return None
    if not str(row_vals[0] or '').strip().startswith('MS'):
        return None
    quotation_no = build_quotation_no(row_vals[0], row_vals[1], row_vals[2], row_vals[3])
    if not quotation_no:
        return None

    year_suffix = safe_str(row_vals[2])
    person_code = normalize_person_code(row_vals[3]) if len(row_vals) > 3 else None
    rec = {
        'quotation_no': quotation_no,
        'source_year': infer_source_year(source_file, year_suffix),
        'quote_date': None,
        'source_file': normalize_master_source_file(source_file),
        'source_sheet': source_sheet,
        'person_code': person_code,
        'doc_type': safe_str(row_vals[4]) or '報價',
        'awarded': '中',
        'site_name': safe_str(row_vals[5]) if len(row_vals) > 5 else None,
        'trade_category': safe_str(row_vals[6]) if len(row_vals) > 6 else None,
        'description': safe_str(row_vals[7]) if len(row_vals) > 7 else None,
        'person_in_charge': safe_str(row_vals[8]) if len(row_vals) > 8 else None,
        'client_name': safe_str(row_vals[9]) if len(row_vals) > 9 else None,
        'quoted_amount': safe_float(row_vals[10]) if len(row_vals) > 10 else None,
        'margin_pct': None,
        'awarded_amount': safe_float(row_vals[10]) if len(row_vals) > 10 else None,
        'contract_days': None,
        'start_date': None,
        'completion_date': None,
        'subcon_type': None,
        'subcon_company': None,
        'subcon_amount': None,
        'profit_amount': None,
        'profit_pct': None,
        'checklist_json': None,
    }
    return enrich_person_fields(rec)


def merge_awarded_record(existing, incoming):
    """合併「中標項目」至同編號的報價主檔，保留報價日期並標記中標。"""
    merged = dict(existing)
    merged['awarded'] = '中'
    if not incoming.get('quote_date') and existing.get('quote_date'):
        merged['quote_date'] = existing['quote_date']
    for field in SYNC_FIELDS:
        if field in ('awarded', 'quote_date'):
            continue
        new_val = incoming.get(field)
        if new_val is None or new_val == '':
            continue
        merged[field] = new_val
    return enrich_person_fields(merged)


def parse_master_row(row_vals, source_file, source_sheet, layout, col_map=None, finance_map=None):
    """依 sheet 表頭版型解析"""
    if source_sheet == '中標項目':
        return parse_awarded_sheet_row(row_vals, source_file, source_sheet)
    if source_sheet == '合約':
        if layout == 'contract_site_amount':
            return parse_contract_site_amount_row(row_vals, source_file, source_sheet)
        return parse_legacy_contract_row(row_vals, source_file, source_sheet)
    if layout == 'modern':
        return parse_modern_row(row_vals, source_file, source_sheet, col_map, finance_map)
    if layout == 'transitional_site':
        return parse_transitional_site_row(row_vals, source_file, source_sheet, col_map)
    if layout == 'transitional_2021':
        return parse_transitional_2021_row(row_vals, source_file, source_sheet, col_map)
    if layout == 'legacy_2018':
        return parse_legacy_2018_row(row_vals, source_file, source_sheet)
    if layout == 'legacy_2017':
        return parse_legacy_2017_row(row_vals, source_file, source_sheet)
    return parse_legacy_quote_row(row_vals, source_file, source_sheet)


def _row_vals(row, min_len=0):
    vals = list(row)
    if min_len and len(vals) < min_len:
        vals.extend([None] * (min_len - len(vals)))
    return vals


def read_master_workbook(filepath):
    """讀取 Master List，回傳 {quotation_no: record}"""
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    records = {}
    errors = []
    rows_read = 0

    for sheet_name in _ordered_master_sheets(wb.sheetnames):
        ws = wb[sheet_name]
        header_found = False
        layout = 'legacy_2017'
        col_map = {}
        finance_map = {}
        for i, row in enumerate(ws.iter_rows(values_only=True), 1):
            row_vals = _row_vals(row)
            if not header_found:
                if _is_header_row(row_vals):
                    header_found = True
                    col_map = map_header_columns(row_vals)
                    from master_finance import map_finance_columns
                    finance_map = map_finance_columns(row_vals)
                    if sheet_name == '合約':
                        layout = detect_contract_layout(row_vals)
                    else:
                        layout = detect_quote_layout(row_vals)
                continue
            if not any(c is not None and str(c).strip() for c in row_vals[:5]):
                continue
            rows_read += 1
            try:
                rec = parse_master_row(row_vals, filepath, sheet_name, layout, col_map, finance_map)
                if not rec:
                    continue
                key = rec['quotation_no']
                if key in records:
                    if sheet_name == '中標項目':
                        records[key] = merge_awarded_record(records[key], rec)
                    elif records[key].get('awarded') == '中':
                        records[key] = merge_awarded_record(rec, records[key])
                    else:
                        errors.append(f'檔內重複: {key}（工作表 {sheet_name} 第 {i} 行）')
                        records[key] = rec
                else:
                    records[key] = rec
            except Exception as e:
                errors.append(f'{sheet_name} 第 {i} 行: {e}')

    wb.close()
    return {
        'records': records,
        'rows_read': rows_read,
        'errors': errors,
        'source_file': normalize_master_source_file(filepath),
    }


def _norm_compare(val):
    if val is None:
        return ''
    if isinstance(val, float):
        return round(val, 2)
    return str(val).strip()


def diff_record(existing, incoming):
    changes = []
    for field in SYNC_FIELDS:
        old = _norm_compare(existing.get(field))
        new = _norm_compare(incoming.get(field))
        if old != new:
            changes.append({'field': field, 'old': existing.get(field), 'new': incoming.get(field)})
    return changes


def preview_master_import(filepath):
    """預覽匯入差異，不寫入 DB"""
    parsed = read_master_workbook(filepath)
    incoming = parsed['records']
    new_rows = []
    updated_rows = []
    unchanged = 0

    for qno, rec in incoming.items():
        existing = db.get_quotation_by_no(qno)
        if not existing:
            new_rows.append(rec)
            continue
        changes = diff_record(existing, rec)
        if changes:
            updated_rows.append({
                'quotation_no': qno,
                'changes': changes,
                'site_name': rec.get('site_name'),
                'awarded': rec.get('awarded'),
            })
        else:
            unchanged += 1

    return {
        'source_file': parsed['source_file'],
        'rows_read': parsed['rows_read'],
        'parse_errors': parsed['errors'],
        'total_in_file': len(incoming),
        'new_count': len(new_rows),
        'updated_count': len(updated_rows),
        'unchanged_count': unchanged,
        'new_sample': new_rows[:20],
        'updated_sample': updated_rows[:20],
    }


def sync_finance_import(filepath):
    """只重匯 Phase 2 財務明細（糧期／分判付款／支票），不覆寫主檔其他欄位"""
    parsed = read_master_workbook(filepath)
    stats = {'finance_rows': 0, 'skipped': 0, 'errors': list(parsed['errors'])}
    for qno, rec in parsed['records'].items():
        finance = rec.get('_finance')
        if not finance or not any(
            finance.get(k) for k in ('summary', 'qs_subcon_lines', 'client_invoices', 'subcon_payments', 'cheques')
        ):
            stats['skipped'] += 1
            continue
        if not db.get_quotation_by_no(qno):
            stats['skipped'] += 1
            continue
        db.replace_quotation_finance(qno, finance)
        db.sync_qs_subcon_registry_fields(qno, finance)
        stats['finance_rows'] += 1
    return stats


def sync_master_import(filepath, preserve_project_link=True):
    """匯入 Master List 至 quotation_registry"""
    parsed = read_master_workbook(filepath)
    incoming = parsed['records']
    stats = {'new': 0, 'updated': 0, 'unchanged': 0, 'errors': list(parsed['errors'])}

    for qno, rec in incoming.items():
        finance = rec.pop('_finance', None)
        existing = db.get_quotation_by_no(qno)
        if existing:
            if preserve_project_link and existing.get('project_id'):
                rec['project_id'] = existing['project_id']
            changes = diff_record(existing, rec)
            if changes:
                db.upsert_quotation_registry(rec)
                stats['updated'] += 1
            else:
                stats['unchanged'] += 1
        else:
            db.upsert_quotation_registry(rec)
            stats['new'] += 1
        if finance:
            db.replace_quotation_finance(qno, finance)

    db.record_master_list_import({
        'source_file': parsed['source_file'],
        'source_year': infer_source_year(filepath),
        'rows_read': parsed['rows_read'],
        'rows_new': stats['new'],
        'rows_updated': stats['updated'],
    })
    return stats


def sync_all_reference(reference_dir):
    """依年份由舊到新匯入資料夾內 Master List"""
    if not os.path.isdir(reference_dir):
        raise FileNotFoundError(reference_dir)
    files = []
    for name in os.listdir(reference_dir):
        if not is_master_list_file(name):
            continue
        m = re.search(r'(20\d{2})', name)
        year = int(m.group(1)) if m else 9999
        files.append((year, name))
    files.sort(key=lambda x: (x[0], x[1]))
    combined = {'files': [], 'totals': {'new': 0, 'updated': 0, 'unchanged': 0}}
    for year, name in files:
        path = os.path.join(reference_dir, name)
        result = sync_master_import(path)
        combined['files'].append({'year': year, 'file': name, **result})
        for k in ('new', 'updated', 'unchanged'):
            combined['totals'][k] += result[k]
    return combined


if __name__ == '__main__':
    db.init_db()
    cmd = (sys.argv[1] if len(sys.argv) > 1 else '').lower()

    if cmd == 'sync-all':
        ref = sys.argv[2] if len(sys.argv) > 2 else os.path.join(
            os.path.dirname(__file__), 'Ref',
        )
        ref = os.path.normpath(ref)
        print(f'[MASTER] 批次匯入: {ref}')
        out = sync_all_reference(ref)
        print(json.dumps(out, ensure_ascii=False, indent=2))
    elif cmd == 'sync-finance' and len(sys.argv) > 2:
        path = sys.argv[2]
        out = sync_finance_import(path)
        print(json.dumps(out, ensure_ascii=False, indent=2, default=str))
    elif cmd in ('preview', 'sync') and len(sys.argv) > 2:
        path = sys.argv[2]
        if cmd == 'preview':
            out = preview_master_import(path)
        else:
            out = sync_master_import(path)
        print(json.dumps(out, ensure_ascii=False, indent=2, default=str))
    else:
        print(__doc__)
