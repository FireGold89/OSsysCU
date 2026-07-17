"""
excel_importer_payment.py — 項目 Payment Status Table 匯入（舊版／對照用）

匯入 MS_Q1241_24 等地盤付款狀況 Excel → projects / subcontractors / payment_records / interim_payments

獨立執行:
  python excel_importer_payment.py [path/to/payment.xlsx]

相容入口仍可用 excel_importer.py（轉呼叫本模組）。
"""
import openpyxl
import os
import re
import sys
from datetime import datetime
import database as db
from sc_ref import derive_parent_sc_no, resolve_contract_amounts


def safe_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() not in ('none', 'nan') else None


def safe_float(val):
    if val is None:
        return 0.0
    try:
        return float(str(val).replace(',', '').replace('$', '').strip())
    except (ValueError, TypeError):
        return 0.0


def safe_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if s in ('-', '—', 'none', 'nan'):
        return None
    for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y']:
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return s if s else None


def _resolve_project_summary_sc(sc_no, company_en, company_zh, description, contract_amt):
    """Project Summary 列：無公司名稱時以工程描述作顯示名（O-005 等代支項）"""
    if company_en or company_zh:
        return company_en, company_zh
    if description:
        return description, company_zh
    if contract_amt and contract_amt > 0:
        return sc_no, company_zh
    return None, None


def read_contract_amount(filepath):
    """從 Excel Summary 工作表讀取承建金額"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    contract_amount = 0
    for sheet_name in wb.sheetnames:
        if 'Summary' not in sheet_name:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True, max_row=20):
            row_vals = [safe_str(c) for c in row]
            for i, v in enumerate(row_vals):
                if v and ('合約金額' in v or '建造金額' in v or '承建金額' in v):
                    for j in range(i + 1, min(i + 5, len(row_vals))):
                        amt = safe_float(row[j] if j < len(row_vals) else None)
                        if amt > 100000:
                            contract_amount = amt
                            return contract_amount
    return contract_amount


def _is_labour_allocation_label(text):
    """Project Summary 右下角 (C1) 財務會作調撥（人工分攤）標籤"""
    if not text:
        return False
    if '(C1)' in text:
        return True
    if '調撥' in text and '人工' in text:
        return True
    if '財務' in text and '人工' in text:
        return True
    return False


def read_labour_allocation(filepath):
    """從 Project Summary 讀取 (C1) 財務會作調撥（人工分攤）金額"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    if 'Project Summary' not in wb.sheetnames:
        return 0
    ws = wb['Project Summary']
    best_amt = 0
    best_prio = -1
    for row in ws.iter_rows(values_only=True):
        rv = list(row)
        for j, cell in enumerate(rv):
            v = safe_str(cell)
            if not _is_labour_allocation_label(v):
                continue
            prio = 2 if '(C1)' in v else 1
            for k in range(j + 1, min(j + 4, len(rv))):
                amt = safe_float(rv[k])
                if amt > 0:
                    if prio > best_prio:
                        best_amt = amt
                        best_prio = prio
                    break
    return best_amt


def sync_labour_allocation_from_excel(filepath, project_id):
    """從 Excel 更新項目人工分攤 (C1)"""
    labour = read_labour_allocation(filepath)
    if labour <= 0:
        return 0
    db.set_project_labour_allocation(project_id, labour)
    print(f'[IMPORT] 人工分攤 (C1): HK${labour:,.2f}')
    return labour


def sync_contract_amount_from_excel(filepath, project_id):
    """若項目承建金額為 0，從 Excel 補充"""
    project = db.get_project(project_id)
    if not project or (project.get('contract_amount') or 0) > 0:
        return False
    amount = read_contract_amount(filepath)
    if amount <= 0:
        return False
    db.update_project(project_id, {
        'project_code': project['project_code'],
        'project_name': project.get('project_name') or project['project_code'],
        'client': project.get('client') or 'MTR',
        'main_contractor': project.get('main_contractor') or '',
        'contract_amount': amount,
        'labour_allocation': project.get('labour_allocation') or 0,
        'start_date': project.get('start_date'),
        'status': project.get('status') or 'Active',
        'notes': project.get('notes') or '',
    })
    print(f"[SYNC] 承建金額已更新: {project['project_code']} = HK${amount:,.0f}")
    return True


def _pct_val(val):
    """Excel 小數或百分比 → 顯示用 0–100"""
    v = safe_float(val)
    if v == 0:
        return None
    return round(v * 100, 2) if abs(v) <= 1.5 else round(v, 2)


def _find_summary_sc_columns(ws, header_row):
    """Summary 分包欄位：SC-004 等表頭列 → {col_index: sc_no}"""
    sc_col_map = {}
    start = max(1, header_row - 20)
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i >= header_row or i < start:
            continue
        rv = list(row)
        found = {}
        for j, c in enumerate(rv):
            s = safe_str(c)
            if s and re.match(r'^SC-\d', s, re.I):
                found[j] = s.upper()
        if len(found) >= 1:
            sc_col_map = found
    return sc_col_map


def _summary_subcon_total_col(sc_col_map):
    """分包總支出欄（SC 欄位之後的第一欄數值總計）"""
    if not sc_col_map:
        return 12
    return max(sc_col_map) + 1


def _find_summary_trade_labels(ws, header_row, sc_col_map):
    """Summary 工種簡稱列（SC 表頭上一行）→ {sc_no: trade_label}"""
    if not sc_col_map:
        return {}
    sc_header_row = None
    start = max(1, header_row - 20)
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i >= header_row or i < start:
            continue
        rv = list(row)
        if any(
            safe_str(rv[j]) and re.match(r'^SC-\d', safe_str(rv[j]), re.I)
            for j in range(len(rv))
        ):
            sc_header_row = i
            break
    if not sc_header_row:
        return {}
    trade_labels = {}
    for try_row in (sc_header_row - 1, sc_header_row - 2):
        if try_row < 1:
            continue
        rv = list(ws.iter_rows(min_row=try_row, max_row=try_row, values_only=True))[0]
        found = 0
        for j, sc_no in sc_col_map.items():
            label = safe_str(rv[j]) if len(rv) > j else None
            if not label or re.match(r'^SC-', label, re.I):
                continue
            if 'SUB-CON' in label.upper() or 'HK$' in label:
                continue
            trade_labels[sc_no] = label
            found += 1
        if found:
            return trade_labels
    return {}


def import_site_ip_period(filepath, project_id):
    """從 Excel Summary 工作表匯入地盤糧期狀況（IP-01…）及分包糧期矩陣"""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = None
    for sn in wb.sheetnames:
        if sn == 'Summary' or sn.lower() == 'summary':
            ws = wb[sn]
            break
    if not ws:
        print('[IMPORT] 找不到 Summary 工作表，跳過糧期匯入')
        return 0

    site_period_text = None
    for row in ws.iter_rows(min_row=1, max_row=20, values_only=True):
        rv = list(row)
        if rv and rv[0] and '工期' in str(rv[0]):
            parts = [safe_str(rv[1]), safe_str(rv[2])]
            site_period_text = ' '.join(p for p in parts if p)
            break

    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        rv = [safe_str(c) for c in row]
        if rv and rv[0] and 'Interim Payment' in rv[0]:
            header_row = i
            break
        if any(v and '糧款期數' in str(v) for v in row if v):
            header_row = i
            break

    if not header_row:
        print('[IMPORT] Summary 無糧款期數表，跳過')
        return 0

    sc_col_map = _find_summary_sc_columns(ws, header_row)
    trade_labels = _find_summary_trade_labels(ws, header_row, sc_col_map)
    total_col = _summary_subcon_total_col(sc_col_map)
    pct_col = total_col + 1

    items = []
    sc_lines = []
    seq = 0
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i <= header_row + 1:
            continue
        rv = list(row)
        ip_no = safe_str(rv[0]) if len(rv) > 0 else None
        if not ip_no or not re.match(r'^IP-\d+', ip_no, re.I):
            continue
        app_amt = safe_float(rv[2]) if len(rv) > 2 else 0
        cert_amt = safe_float(rv[4]) if len(rv) > 4 else 0
        if not app_amt and not cert_amt:
            break
        seq += 1
        ip_key = ip_no.upper()
        subcon_total = safe_float(rv[total_col]) if len(rv) > total_col else 0
        if not subcon_total and sc_col_map:
            subcon_total = sum(
                safe_float(rv[j]) for j in sc_col_map if len(rv) > j
            )
        items.append({
            'ip_no': ip_key,
            'seq_no': seq,
            'applied_date': safe_date(rv[1]) if len(rv) > 1 else None,
            'application_amount': app_amt,
            'certified_income': cert_amt,
            'certificate_date': safe_date(rv[6]) if len(rv) > 6 else None,
            'subcon_paid': subcon_total,
            'subcon_paid_pct': _pct_val(rv[pct_col]) if len(rv) > pct_col else None,
            'subcon_cert_date': None,
        })
        for j, sc_no in sc_col_map.items():
            amt = safe_float(rv[j]) if len(rv) > j else 0
            if amt:
                sc_lines.append({'ip_no': ip_key, 'sc_no': sc_no, 'amount': amt})

    totals = {'total_income': 0, 'total_expenditure': 0, 'advance': 0}
    for row in ws.iter_rows(min_row=header_row, max_row=header_row + 25, values_only=True):
        rv = [safe_str(c) if c is not None else None for c in row]
        for j, v in enumerate(rv):
            if not v:
                continue
            nxt = safe_float(rv[j + 1]) if j + 1 < len(rv) else 0
            if '總收入' in v and nxt:
                totals['total_income'] = nxt
            elif '總支出' in v and nxt:
                totals['total_expenditure'] = nxt
            elif '墊支' in v and nxt:
                totals['advance'] = nxt

    if not items:
        return 0

    db.replace_interim_payments(project_id, items, {
        'site_period_text': site_period_text,
        'ip_total_income': totals['total_income'],
        'ip_total_expenditure': totals['total_expenditure'],
        'ip_advance': totals['advance'],
    }, sc_lines=sc_lines)
    if trade_labels:
        db.set_subcontractor_trade_labels(project_id, trade_labels)
    sc_info = f'、{len(sc_lines)} 筆分包明細' if sc_lines else ''
    trade_info = f'、{len(trade_labels)} 個工種標籤' if trade_labels else ''
    print(f'[IMPORT] 已匯入 {len(items)} 期糧款記錄（地盤糧期狀況{sc_info}{trade_info}）')
    return len(items)


def import_excel(filepath, project_code=None):
    """匯入整個Excel文件到資料庫"""
    print(f"\n[IMPORT] 開始匯入: {filepath}")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets = wb.sheetnames
    print(f"[IMPORT] 工作表: {sheets}")

    # ─── 建立/取得項目 ───────────────────────────────────────────
    # 從第一個工作表嘗試讀取項目名稱
    ws_first = wb[sheets[0]]
    project_name = None
    for row in ws_first.iter_rows(values_only=True, max_row=5):
        for cell in row:
            val = safe_str(cell)
            if val and 'MTR' in val and len(val) > 20:
                project_name = val
                break
        if project_name:
            break

    if not project_code:
        # 從文件名提取項目代碼
        basename = os.path.basename(filepath)
        # 嘗試提取 MS_Q1241_24 格式
        import re
        m = re.search(r'(MS_Q\d+_\d+|Q\d+)', basename, re.IGNORECASE)
        project_code = m.group(1) if m else os.path.splitext(basename)[0][:20]

    # 從 Project Summary / Summary 讀取合約金額
    contract_amount = 0
    main_contractor = None
    client = None

    for sheet_name in sheets:
        if 'Summary' in sheet_name or 'summary' in sheet_name.lower():
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True, max_row=15):
                row_vals = [safe_str(c) for c in row]
                # 搜尋主合約資料
                for i, v in enumerate(row_vals):
                    if v and ('總承建' in v or 'Main Contractor' in v.lower() if v else False):
                        if i + 1 < len(row_vals) and row_vals[i + 1]:
                            main_contractor = row_vals[i + 1]
                    if v and ('合約金額' in v or '建造金額' in v or '承建金額' in v or 'Contract' in v.lower() if v else False):
                        for j in range(i + 1, min(i + 5, len(row_vals))):
                            amt = safe_float(row[j] if j < len(row) else None)
                            if amt > 100000:
                                contract_amount = amt
                                break
            break

    # 嘗試從 Project Summary 工作表讀取
    if 'Project Summary' in sheets:
        ws_ps = wb['Project Summary']
        for i, row in enumerate(ws_ps.iter_rows(values_only=True, max_row=8)):
            row_vals = list(row)
            for j, cell in enumerate(row_vals):
                v = safe_str(cell)
                if v and ('總承建' in v or 'Main Contractor' in v.lower() if v else False):
                    if j + 1 < len(row_vals):
                        main_contractor = safe_str(row_vals[j + 1])
                if v and ('建造金額' in v or '承建金額' in v or '工程金額' in v):
                    for k in range(j + 1, min(j + 5, len(row_vals))):
                        amt = safe_float(row_vals[k] if k < len(row_vals) else None)
                        if amt > 100000:
                            contract_amount = amt
                            break

    # 建立項目
    existing_projects = db.get_all_projects()
    project_id = None
    for p in existing_projects:
        if p['project_code'] == project_code:
            project_id = p['id']
            print(f"[IMPORT] 項目已存在 (ID={project_id}): {project_code}")
            break

    if project_id and contract_amount > 0:
        existing = db.get_project(project_id)
        if existing and (not existing.get('contract_amount') or existing['contract_amount'] <= 0):
            db.update_project(project_id, {
                'project_code': project_code,
                'project_name': project_name or existing.get('project_name') or project_code,
                'project_name_en': project_name or existing.get('project_name_en') or existing.get('project_name') or '',
                'project_name_zh': existing.get('project_name_zh') or '',
                'client': existing.get('client') or client or 'MTR',
                'main_contractor': main_contractor or existing.get('main_contractor') or '',
                'contract_amount': contract_amount,
                'labour_allocation': existing.get('labour_allocation') or 0,
                'start_date': existing.get('start_date'),
                'status': existing.get('status') or 'Active',
                'notes': existing.get('notes') or '',
            })
            print(f"[IMPORT] 已更新承建金額: HK${contract_amount:,.0f}")

    if not project_id:
        project_id = db.create_project({
            'project_code': project_code,
            'project_name': project_name or project_code,
            'project_name_en': project_name or project_code,
            'project_name_zh': '',
            'client': client or 'MTR',
            'main_contractor': main_contractor or 'Mepork Engineering Services Limited',
            'contract_amount': contract_amount,
            'labour_allocation': read_labour_allocation(filepath),
            'start_date': None,
            'status': 'Active',
            'notes': f'從 {os.path.basename(filepath)} 匯入'
        })
        print(f"[IMPORT] 已建立項目 (ID={project_id}): {project_code}")

    # ─── 匯入 Project Summary 工作表（分判商清單）───────────────
    sc_map = {}  # sc_no -> sc_id

    if 'Project Summary' in sheets:
        ws_ps = wb['Project Summary']
        print(f"\n[IMPORT] 匯入分判商清單...")
        sc_count = 0
        header_row = None

        for i, row in enumerate(ws_ps.iter_rows(values_only=True)):
            row_vals = list(row)
            # 找標頭行
            if header_row is None:
                row_strs = [safe_str(c) for c in row_vals]
                if any(v and ('No.' in v or 'SC' in v or 'Company' in v.lower()) for v in row_strs if v):
                    header_row = i
                    continue

            if header_row is not None and i > header_row:
                # 嘗試讀取分判商資料
                # 欄位: No, SC_No, Quotation_No, Company_EN, Company_ZH, Description, ?, Contract_Amount, ...
                if len(row_vals) < 4:
                    continue
                sc_no = safe_str(row_vals[1])
                if not sc_no or sc_no.startswith('('):
                    continue
                if sc_no in ('M-0XX', 'SC-0XX', 'O-0XX'):
                    continue

                company_en = safe_str(row_vals[3]) if len(row_vals) > 3 else None
                company_zh = safe_str(row_vals[4]) if len(row_vals) > 4 else None
                description = safe_str(row_vals[5]) if len(row_vals) > 5 else None
                contract_charge = safe_str(row_vals[6]) if len(row_vals) > 6 else None
                # H=Contract Sum, I=VO, J=Revised Contract Sum (J = H + VO)
                h_raw = safe_float(row_vals[7]) if len(row_vals) > 7 else 0
                vo_raw = safe_float(row_vals[8]) if len(row_vals) > 8 else 0
                j_raw = safe_float(row_vals[9]) if len(row_vals) > 9 else 0
                contract_sum, vo_amount, contract_amt = resolve_contract_amounts(
                    h_raw, vo_raw, j_raw if j_raw > 0 else None)
                is_excluded = 1 if contract_charge == '*' else 0
                quotation_no = safe_str(row_vals[2]) if len(row_vals) > 2 else None

                # 狀態欄：O 欄 = OA 日期（放入 OA 系統日期，非報價單日期）
                oa_status = safe_str(row_vals[11]) if len(row_vals) > 11 else None
                oa_ref = safe_str(row_vals[12]) if len(row_vals) > 12 else None
                oa_no = safe_str(row_vals[13]) if len(row_vals) > 13 else None
                oa_date = safe_date(row_vals[14]) if len(row_vals) > 14 else None
                quotation_saved = safe_str(row_vals[15]) if len(row_vals) > 15 else None
                payment_note = safe_str(row_vals[10]) if len(row_vals) > 10 else None

                company_en, company_zh = _resolve_project_summary_sc(
                    sc_no, company_en, company_zh, description, contract_amt)
                if not company_en and not company_zh:
                    continue

                sc_id = db.upsert_subcontractor({
                    'project_id': project_id,
                    'sc_no': sc_no,
                    'parent_sc_no': derive_parent_sc_no(sc_no),
                    'quotation_no': quotation_no,
                    'company_name_en': company_en,
                    'company_name_zh': company_zh,
                    'description': description,
                    'contract_sum': contract_sum,
                    'vo_amount': vo_amount,
                    'contract_amount': contract_amt,
                    'payment_note': payment_note,
                    'oa_status': oa_status,
                    'oa_ref': oa_ref,
                    'oa_no': oa_no,
                    'quotation_saved': quotation_saved,
                    'quotation_date': None,
                    'oa_date': oa_date,
                    'is_excluded': is_excluded,
                })
                sc_map[sc_no] = sc_id
                sc_count += 1
                print(f"  [{sc_count}] {sc_no}: {company_en}")

        print(f"[IMPORT] 已匯入 {sc_count} 個分判商")

    # ─── 匯入付款記錄工作表 ────────────────────────────────────
    payment_sheet = None
    for sname in sheets:
        sname_lower = sname.lower()
        # 匹配中文或英文付款表關鍵詞
        if any(kw in sname for kw in ['分判', '付款', '工程']) or 'payment' in sname_lower:
            payment_sheet = sname
            break
    # 如果還是找不到，選擇第2個工作表（通常是付款表）
    if not payment_sheet and len(sheets) >= 2:
        payment_sheet = sheets[1]
    print(f"[IMPORT] 付款工作表: {payment_sheet}")

    if payment_sheet:
        ws_pay = wb[payment_sheet]
        print(f"\n[IMPORT] 匯入付款記錄 (工作表: {payment_sheet})...")
        header_row = None
        pay_count = 0

        for i, row in enumerate(ws_pay.iter_rows(values_only=True)):
            row_vals = list(row)
            # 找標頭行（含 Invoice Date, Invoice No. 等欄位名）
            if header_row is None:
                row_strs = [safe_str(c) for c in row_vals]
                if any(v and 'Invoice' in v for v in row_strs if v):
                    header_row = i
                    # 記錄欄位索引
                    col_map = {}
                    for j, v in enumerate(row_strs):
                        if v:
                            v_lower = v.lower().strip()
                            if 'invoice date' in v_lower:
                                col_map['invoice_date'] = j
                            elif 'invoice no' in v_lower:
                                col_map['invoice_no'] = j
                            elif 'quotation' in v_lower:
                                col_map['quotation_no'] = j
                            elif "sub-contractor" in v_lower and 'no' in v_lower:
                                col_map['sc_no'] = j
                            elif 'company name' in v_lower and 'chinese' in v_lower:
                                col_map['company_name_zh'] = j
                            elif 'company name' in v_lower:
                                col_map['company_name_en'] = j
                            elif 'description' in v_lower:
                                col_map['description'] = j
                            elif 'contract amount' in v_lower:
                                col_map['contract_amount'] = j
                            elif 'paid amount' in v_lower:
                                col_map['paid_amount'] = j
                            elif 'remainder' in v_lower:
                                col_map['remainder_amount'] = j
                            elif 'oa' in v_lower and '狀' in v_lower:
                                col_map['oa_ref'] = j
                            elif 'oa' in v_lower and '編' in v_lower:
                                col_map['oa_no'] = j
                            elif 'mc ip' in v_lower:
                                col_map['mc_ip_no'] = j
                            elif 'b/c' in v_lower or 'bc' in v_lower:
                                col_map['bc_to_sub'] = j
                            elif 'sub-ip' in v_lower or 'sub ip' in v_lower:
                                col_map['sub_ip_no'] = j
                            elif 'remark' in v_lower or 'remrk' in v_lower:
                                col_map['remark'] = j
                    print(f"  [欄位映射] {col_map}")
                    continue

            if header_row is not None and i > header_row:
                if len(row_vals) < 6:
                    continue

                seq_no = safe_str(row_vals[0])
                if not seq_no or seq_no in ('No.', 'Link'):
                    continue

                # 讀取各欄位
                def get_col(key, default=None):
                    idx = col_map.get(key)
                    return row_vals[idx] if idx is not None and idx < len(row_vals) else default

                invoice_date = safe_date(get_col('invoice_date'))
                invoice_no = safe_str(get_col('invoice_no'))
                quotation_no = safe_str(get_col('quotation_no'))
                sc_no = safe_str(get_col('sc_no'))
                company_en = safe_str(get_col('company_name_en'))
                company_zh = safe_str(get_col('company_name_zh'))
                description = safe_str(get_col('description'))
                contract_amt = safe_float(get_col('contract_amount'))
                paid_amt = safe_float(get_col('paid_amount'))
                remainder = safe_float(get_col('remainder_amount'))
                oa_ref = safe_str(get_col('oa_ref'))
                oa_no = safe_str(get_col('oa_no'))
                mc_ip_no = safe_str(get_col('mc_ip_no'))
                bc_to_sub = safe_str(get_col('bc_to_sub'))
                sub_ip_no = safe_str(get_col('sub_ip_no'))
                remark = safe_str(get_col('remark'))

                # 必須有分判商編號或公司名
                if not sc_no and not company_en:
                    continue

                sc_id = sc_map.get(sc_no) if sc_no else None

                pay_id = db.create_payment({
                    'project_id': project_id,
                    'sc_id': sc_id,
                    'seq_no': seq_no,
                    'invoice_date': invoice_date,
                    'invoice_no': invoice_no,
                    'quotation_no': quotation_no,
                    'sc_no': sc_no,
                    'company_name_en': company_en,
                    'company_name_zh': company_zh,
                    'description': description,
                    'contract_amount': contract_amt,
                    'paid_amount': paid_amt,
                    'remainder_amount': remainder,
                    'oa_ref': oa_ref,
                    'oa_no': oa_no,
                    'mc_ip_no': mc_ip_no,
                    'bc_to_sub': bc_to_sub,
                    'sub_ip_no': sub_ip_no,
                    'remark': remark,
                    'pdf_path': None,
                    'ocr_status': None,
                })
                pay_count += 1
                print(f"  [{pay_count}] #{seq_no} {sc_no} {company_en}: 已付 ${paid_amt:,.0f}")

        print(f"[IMPORT] 已匯入 {pay_count} 條付款記錄")

    import_site_ip_period(filepath, project_id)
    sync_labour_allocation_from_excel(filepath, project_id)

    print(f"\n[IMPORT] 完成! 項目ID: {project_id}")
    return project_id


def sync_excel_data(filepath, project_id=None):
    """同步 Excel → 合同項目、付款、糧期（保留 OCR 報價日期/PDF）"""
    print(f"\n[SYNC] 開始同步: {filepath}")
    db.init_db()

    if not os.path.exists(filepath):
        raise FileNotFoundError(filepath)

    if not project_id:
        projects = db.get_all_projects()
        if not projects:
            print('[SYNC] 無項目，改為完整匯入')
            return import_excel(filepath)
        project_id = projects[0]['id']
        print(f"[SYNC] 使用項目 ID={project_id} ({projects[0].get('project_code')})")

    sync_contract_amount_from_excel(filepath, project_id)
    sync_labour_allocation_from_excel(filepath, project_id)
    db.replace_payments_for_project(project_id)
    db.replace_subcontractors_for_project(project_id)
    print('[SYNC] 已清除舊付款及分判商，重新匯入...')

    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets = wb.sheetnames
    sc_map = {}

    if 'Project Summary' in sheets:
        ws_ps = wb['Project Summary']
        print('[SYNC] 同步合同項目...')
        sc_count = 0
        header_row = None
        for i, row in enumerate(ws_ps.iter_rows(values_only=True)):
            row_vals = list(row)
            if header_row is None:
                row_strs = [safe_str(c) for c in row_vals]
                if any(v and ('No.' in v or 'SC' in v or 'Company' in v.lower()) for v in row_strs if v):
                    header_row = i
                    continue
            if header_row is not None and i > header_row:
                if len(row_vals) < 4:
                    continue
                sc_no = safe_str(row_vals[1])
                if not sc_no or sc_no.startswith('(') or sc_no in ('M-0XX', 'SC-0XX', 'O-0XX'):
                    continue
                company_en = safe_str(row_vals[3]) if len(row_vals) > 3 else None
                company_zh = safe_str(row_vals[4]) if len(row_vals) > 4 else None
                description = safe_str(row_vals[5]) if len(row_vals) > 5 else None
                contract_charge = safe_str(row_vals[6]) if len(row_vals) > 6 else None
                h_raw = safe_float(row_vals[7]) if len(row_vals) > 7 else 0
                vo_raw = safe_float(row_vals[8]) if len(row_vals) > 8 else 0
                j_raw = safe_float(row_vals[9]) if len(row_vals) > 9 else 0
                contract_sum, vo_amount, contract_amt = resolve_contract_amounts(
                    h_raw, vo_raw, j_raw if j_raw > 0 else None)
                is_excluded = 1 if contract_charge == '*' else 0
                quotation_no = safe_str(row_vals[2]) if len(row_vals) > 2 else None
                oa_status = safe_str(row_vals[11]) if len(row_vals) > 11 else None
                oa_ref = safe_str(row_vals[12]) if len(row_vals) > 12 else None
                oa_no = safe_str(row_vals[13]) if len(row_vals) > 13 else None
                oa_date = safe_date(row_vals[14]) if len(row_vals) > 14 else None
                quotation_saved = safe_str(row_vals[15]) if len(row_vals) > 15 else None
                payment_note = safe_str(row_vals[10]) if len(row_vals) > 10 else None
                company_en, company_zh = _resolve_project_summary_sc(
                    sc_no, company_en, company_zh, description, contract_amt)
                if not company_en and not company_zh:
                    continue
                sc_id = db.upsert_subcontractor({
                    'project_id': project_id,
                    'sc_no': sc_no,
                    'parent_sc_no': derive_parent_sc_no(sc_no),
                    'quotation_no': quotation_no,
                    'company_name_en': company_en,
                    'company_name_zh': company_zh,
                    'description': description,
                    'contract_sum': contract_sum,
                    'vo_amount': vo_amount,
                    'contract_amount': contract_amt,
                    'payment_note': payment_note,
                    'oa_status': oa_status,
                    'oa_ref': oa_ref,
                    'oa_no': oa_no,
                    'quotation_saved': quotation_saved,
                    'quotation_date': None,
                    'oa_date': oa_date,
                    'is_excluded': is_excluded,
                })
                sc_map[sc_no] = sc_id
                sc_count += 1
        print(f'[SYNC] 已同步 {sc_count} 個合同項目')

    payment_sheet = None
    for sname in sheets:
        if any(kw in sname for kw in ['分判', '付款', '工程']) or 'payment' in sname.lower():
            payment_sheet = sname
            break
    if not payment_sheet and len(sheets) >= 2:
        payment_sheet = sheets[1]

    pay_count = 0
    if payment_sheet:
        ws_pay = wb[payment_sheet]
        print(f'[SYNC] 同步付款記錄 ({payment_sheet})...')
        header_row = None
        col_map = {}
        for i, row in enumerate(ws_pay.iter_rows(values_only=True)):
            row_vals = list(row)
            if header_row is None:
                row_strs = [safe_str(c) for c in row_vals]
                if any(v and 'Invoice' in v for v in row_strs if v):
                    header_row = i
                    for j, v in enumerate(row_strs):
                        if not v:
                            continue
                        v_lower = v.lower().strip()
                        if 'invoice date' in v_lower:
                            col_map['invoice_date'] = j
                        elif 'invoice no' in v_lower:
                            col_map['invoice_no'] = j
                        elif 'quotation' in v_lower:
                            col_map['quotation_no'] = j
                        elif "sub-contractor" in v_lower and 'no' in v_lower:
                            col_map['sc_no'] = j
                        elif 'company name' in v_lower and 'chinese' in v_lower:
                            col_map['company_name_zh'] = j
                        elif 'company name' in v_lower:
                            col_map['company_name_en'] = j
                        elif 'description' in v_lower:
                            col_map['description'] = j
                        elif 'contract amount' in v_lower:
                            col_map['contract_amount'] = j
                        elif 'paid amount' in v_lower:
                            col_map['paid_amount'] = j
                        elif 'remainder' in v_lower:
                            col_map['remainder_amount'] = j
                        elif 'oa' in v_lower and '狀' in v_lower:
                            col_map['oa_ref'] = j
                        elif 'oa' in v_lower and '編' in v_lower:
                            col_map['oa_no'] = j
                        elif 'mc ip' in v_lower:
                            col_map['mc_ip_no'] = j
                        elif 'b/c' in v_lower or 'bc' in v_lower:
                            col_map['bc_to_sub'] = j
                        elif 'sub-ip' in v_lower or 'sub ip' in v_lower:
                            col_map['sub_ip_no'] = j
                        elif 'remark' in v_lower or 'remrk' in v_lower:
                            col_map['remark'] = j
                    continue
            if header_row is not None and i > header_row:
                if len(row_vals) < 6:
                    continue
                seq_no = safe_str(row_vals[0])
                if not seq_no or seq_no in ('No.', 'Link'):
                    continue

                def get_col(key, default=None):
                    idx = col_map.get(key)
                    return row_vals[idx] if idx is not None and idx < len(row_vals) else default

                sc_no = safe_str(get_col('sc_no'))
                company_en = safe_str(get_col('company_name_en'))
                if not sc_no and not company_en:
                    continue
                sc_id = sc_map.get(sc_no) if sc_no else None
                db.create_payment({
                    'project_id': project_id,
                    'sc_id': sc_id,
                    'seq_no': seq_no,
                    'invoice_date': safe_date(get_col('invoice_date')),
                    'invoice_no': safe_str(get_col('invoice_no')),
                    'quotation_no': safe_str(get_col('quotation_no')),
                    'sc_no': sc_no,
                    'company_name_en': company_en,
                    'company_name_zh': safe_str(get_col('company_name_zh')),
                    'description': safe_str(get_col('description')),
                    'contract_amount': safe_float(get_col('contract_amount')),
                    'paid_amount': safe_float(get_col('paid_amount')),
                    'remainder_amount': safe_float(get_col('remainder_amount')),
                    'oa_ref': safe_str(get_col('oa_ref')),
                    'oa_no': safe_str(get_col('oa_no')),
                    'mc_ip_no': safe_str(get_col('mc_ip_no')),
                    'bc_to_sub': safe_str(get_col('bc_to_sub')),
                    'sub_ip_no': safe_str(get_col('sub_ip_no')),
                    'remark': safe_str(get_col('remark')),
                    'pdf_path': None,
                    'ocr_status': None,
                })
                pay_count += 1
        print(f'[SYNC] 已匯入 {pay_count} 條付款記錄')

    ip_count = import_site_ip_period(filepath, project_id)
    print(f'[SYNC] 完成! 項目ID={project_id}, 付款={pay_count}, 糧期={ip_count}')
    return project_id


if __name__ == '__main__':
    db.init_db()
    excel_path = os.path.join(
        os.path.dirname(__file__),
        'MS_Q1241_24 - Main contract Works Payment Status Table - R5.xlsx',
    )
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    if os.path.exists(excel_path):
        project_id = import_excel(excel_path)
        print(f'\n[PAYMENT IMPORT] 完成，項目 ID: {project_id}')
    else:
        print(f'[PAYMENT IMPORT] 找不到文件: {excel_path}')
