"""
master_trade_ref.py — Master List 工作範疇選項

從 Ref/YYYY Quotation & Contract number.xlsx 提取：
  - I 欄 工作範疇     → field_type=scope（大類別依 I→工程類別配對推斷）
  - 工程類別分類      → field_type=override（優先「分類清單」+ 資料中額外出現者）
"""
from __future__ import annotations

import os
import re
from collections import Counter, defaultdict

import openpyxl

from config import BASE_DIR
from master_list_importer import _is_header_row, detect_quote_layout

_REF_DIR = os.path.join(BASE_DIR, 'Ref')
_DATA_SHEETS = ('報價', '標書')
_SHEET_FENLEI = '分類清單'
FIELD_SCOPE = 'scope'
FIELD_OVERRIDE = 'override'
EXTRA_CATALOG_GROUP = '資料補充'
UNGROUPED_LABEL = '未分類'


def _s(val):
    if val is None:
        return ''
    return str(val).strip()


def _is_trade_text(val) -> bool:
    s = _s(val)
    if not s or len(s) < 2:
        return False
    if re.match(r'^[\d.,]+$', s):
        return False
    if re.match(r'^\d{4,}$', s.replace(',', '')):
        return False
    return True


def find_master_ref_files(ref_dir: str | None = None) -> list[str]:
    root = ref_dir or _REF_DIR
    if not os.path.isdir(root):
        return []
    files = []
    for name in os.listdir(root):
        if name.startswith('~$') or not name.lower().endswith('.xlsx'):
            continue
        nl = name.lower()
        if 'quotation' in nl and 'contract number' in nl:
            files.append(os.path.join(root, name))
    files.sort(key=lambda p: (re.search(r'(20\d{2})', os.path.basename(p)) or ['0'])[0], reverse=True)
    return files


def find_master_ref_file(ref_dir: str | None = None) -> str | None:
    files = find_master_ref_files(ref_dir)
    return files[0] if files else None


def parse_fenlei_from_workbook(filepath: str) -> list[dict]:
    """「分類清單」sheet → 工程類別分類（含一級代碼）"""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        if _SHEET_FENLEI not in wb.sheetnames:
            return []
        ws = wb[_SHEET_FENLEI]
        records = []
        l1_code = None
        sort_order = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            cells = list(row) + [None] * 4
            if _s(cells[0]):
                l1_code = _s(cells[0])
            name_zh = _s(cells[1])
            if not name_zh or not l1_code:
                continue
            records.append({
                'name_zh': name_zh,
                'l1_code': l1_code,
                'sort_order': sort_order,
                'source_file': os.path.basename(filepath),
            })
            sort_order += 1
        return records
    finally:
        wb.close()


def merge_fenlei_catalog(files: list[str]) -> dict[str, dict]:
    """合併各年份分類清單；同名以較新年份的一級／排序為準"""
    catalog: dict[str, dict] = {}
    for fp in reversed(files):
        for item in parse_fenlei_from_workbook(fp):
            catalog[item['name_zh']] = item
    return catalog


def _scan_workbook_trade_values(filepath: str) -> tuple[dict[str, int], dict[str, int], dict[tuple[str, str], int]]:
    scope_counts: dict[str, int] = {}
    override_counts: dict[str, int] = {}
    pair_counts: dict[tuple[str, str], int] = {}
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            if sheet_name not in _DATA_SHEETS:
                continue
            ws = wb[sheet_name]
            layout = None
            for row in ws.iter_rows(values_only=True):
                rv = list(row)
                if _is_header_row(rv):
                    layout = detect_quote_layout(rv)
                    continue
                if layout != 'modern' or len(rv) < 10:
                    continue
                v8 = _s(rv[8]) if len(rv) > 8 else ''
                v9 = _s(rv[9]) if len(rv) > 9 else ''
                if _is_trade_text(v8):
                    scope_counts[v8] = scope_counts.get(v8, 0) + 1
                if _is_trade_text(v9):
                    override_counts[v9] = override_counts.get(v9, 0) + 1
                if _is_trade_text(v8) and _is_trade_text(v9):
                    key = (v8, v9)
                    pair_counts[key] = pair_counts.get(key, 0) + 1
    finally:
        wb.close()
    return scope_counts, override_counts, pair_counts


def _infer_scope_group_name(scope: str, pair_map: dict[str, Counter]) -> str | None:
    """依 Excel I→工程類別同列配對，取最常見工程類別作為 I 欄大類別"""
    counter = pair_map.get(scope)
    if counter:
        return counter.most_common(1)[0][0]
    return None


def _build_override_records(
    catalog: dict[str, dict],
    override_counts: dict[str, int],
    source_file: str,
) -> list[dict]:
    records: list[dict] = []
    sort_order = 0
    catalog_names = set(catalog.keys())

    for item in sorted(catalog.values(), key=lambda x: x['sort_order']):
        name = item['name_zh']
        records.append({
            'field_type': FIELD_OVERRIDE,
            'name_zh': name,
            'group_name': item.get('l1_code'),
            'use_count': override_counts.get(name, 0),
            'sort_order': sort_order,
            'source_file': item.get('source_file') or source_file,
        })
        sort_order += 1

    extras = set(override_counts.keys()) - catalog_names
    for name in sorted(extras, key=lambda x: (-override_counts.get(x, 0), x)):
        records.append({
            'field_type': FIELD_OVERRIDE,
            'name_zh': name,
            'group_name': EXTRA_CATALOG_GROUP,
            'use_count': override_counts[name],
            'sort_order': sort_order,
            'source_file': source_file,
        })
        sort_order += 1

    if not catalog and override_counts:
        for name, count in sorted(override_counts.items(), key=lambda x: (-x[1], x[0])):
            records.append({
                'field_type': FIELD_OVERRIDE,
                'name_zh': name,
                'group_name': EXTRA_CATALOG_GROUP,
                'use_count': count,
                'sort_order': sort_order,
                'source_file': source_file,
            })
            sort_order += 1

    return records


def parse_trade_options_from_ref(ref_dir: str | None = None) -> list[dict]:
    """合併 Ref Master List：I 欄細項 + 工程類別分類（分類清單 + 資料補充）"""
    files = find_master_ref_files(ref_dir)
    if not files:
        raise FileNotFoundError(f'找不到 Master List 參考檔: {ref_dir or _REF_DIR}')

    scope_counts: dict[str, int] = {}
    override_counts: dict[str, int] = {}
    pair_counts: dict[tuple[str, str], int] = {}
    source_names = []
    for fp in files:
        s8, s9, pairs = _scan_workbook_trade_values(fp)
        for k, v in s8.items():
            scope_counts[k] = scope_counts.get(k, 0) + v
        for k, v in s9.items():
            override_counts[k] = override_counts.get(k, 0) + v
        for k, v in pairs.items():
            pair_counts[k] = pair_counts.get(k, 0) + v
        source_names.append(os.path.basename(fp))

    catalog = merge_fenlei_catalog(files)
    pair_map: dict[str, Counter] = defaultdict(Counter)
    for (scope, override), count in pair_counts.items():
        pair_map[scope][override] += count

    source_file = source_names[0] if len(source_names) == 1 else f'{len(source_names)} files'
    records: list[dict] = []
    sort_order = 0
    for name, count in sorted(scope_counts.items(), key=lambda x: (-x[1], x[0])):
        group_name = _infer_scope_group_name(name, pair_map)
        records.append({
            'field_type': FIELD_SCOPE,
            'name_zh': name,
            'group_name': group_name,
            'use_count': count,
            'sort_order': sort_order,
            'source_file': source_file,
        })
        sort_order += 1

    records.extend(_build_override_records(catalog, override_counts, source_file))
    return records


def resolve_trade_category(scope=None, override=None, fallback=None):
    ov = _s(override)
    sc = _s(scope)
    if ov:
        return ov
    if sc:
        return sc
    fb = _s(fallback)
    return fb or None


def ref_status(ref_dir: str | None = None) -> dict:
    files = find_master_ref_files(ref_dir)
    latest = files[0] if files else None
    catalog = merge_fenlei_catalog(files) if files else {}
    return {
        'ok': bool(files),
        'ref_path': latest,
        'ref_file': os.path.basename(latest) if latest else None,
        'ref_dir': ref_dir or _REF_DIR,
        'ref_file_count': len(files),
        'fenlei_count': len(catalog),
    }
