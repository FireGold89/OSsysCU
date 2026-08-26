"""
engineering_category_ref.py — 美博工程標準分類表（R1）

Ref/Mepork_Engineering_Standard_Category_Table R1.xlsx
用於工程項目（projects）一級／二級分類。
"""
from __future__ import annotations

import os
import re
from datetime import datetime

import openpyxl

from config import BASE_DIR

DEFAULT_REF_PATH = os.path.join(
    BASE_DIR, 'Ref', 'Mepork_Engineering_Standard_Category_Table R1.xlsx',
)
_REF_NAME_HINTS = ('Engineering_Standard_Category', '工程標準分類', '美博工程標準分類')

# 關鍵字 → l2_code（越前越優先）
CLASSIFY_RULES: list[tuple[str, tuple[str, ...]]] = [
    ('4.3', (
        'wr2', 'fs251', '洗水缸', '法定', '強制驗樓', 'mbis', 'statutory inspection',
        '五交年', '五(交)年', '驗樓',
    )),
    ('4.2', (
        '保養合約', '定期保養', '定期清洗', '閘機保養', '水泵系統測試', 'periodic maintenance',
        '設施保養', '保養及改善',
    )),
    ('4.1', (
        '區域保養', '屋邨保養', 'asd', '日常修葺', 'handyman', '零星', '園藝', '休憩區', '休憩設施',
    )),
    ('2.4', (
        'cctv', '閉路電視', '門禁', '對講', '弱電', 'elv', '保安系統', 'security system',
    )),
    ('2.3', (
        '水喉', '渠務', '水泵', 'plumbing', 'drainage', '主渠', '水管', '生鐵渠', 'pvc',
        '給排水', '供水', '排水', '地下排水',
    )),
    ('2.2', (
        '冷氣', '通風', 'mvac', 'hvac', '風喉', '壓縮機', '抽氣', '鮮風',
    )),
    ('2.1', (
        '電力', '照明', '配電', 'electrical', '電箱', 'busduct', '上升總線', '供電',
        '機電', '升降機', '信號機房', '機房', '電線井', 'eva',
    )),
    ('1.4', (
        '鐵器', '木器', '油漆', '玻璃', '防煙門', '欄杆', '扶手', 'joinery', 'metal',
        '鐵閘', '門鎖', '傢俬', '天棚', 'awning', '簷蓬', '簷篷', 'platform awning', '圍欄',
    )),
    ('1.3', (
        '防水', '石屎', '剝落', 'spalling', 'concrete repair', '敲鑿', '防鏽', '還原', '地基',
    )),
    ('1.2', (
        '外牆', '翻新', '維修令', 'bd order', 'refurbishment', '大維修', '磁磚', '石材',
        '物業維修', '維修工程', '改善工程', '空置單位',
    )),
    ('1.1', (
        '加建', '改建', '加設', 'a&a', '打拆', '間隔', '護墻', '護牆', '防護網', '結構開孔',
        'builder', '出口', '隧道', '更換及加設',
    )),
    ('3.1', (
        '裝修', 'fit-out', 'fitness', '健身室', 'fitness room', '假天花', '間隔牆',
        'decoration', '面飾', 'comprehensive fit', '休息室', '酒吧', '廁所',
    )),
]

_PLACEHOLDER_WORK_TYPES = frozenset({'跟分頁既分類', '跟分頁的分類', ''})


def _s(val):
    if val is None:
        return ''
    return str(val).strip()


def _split_bilingual(text: str) -> tuple[str, str]:
    """「中文 / (English)」→ (zh, en)"""
    t = _s(text).replace('\n', ' / ')
    if not t:
        return '', ''
    m = re.match(r'^(.+?)\s*/\s*\(([^)]+)\)\s*$', t)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    m = re.match(r'^(.+?)\s*/\s*(.+)$', t)
    if m and re.search(r'[A-Za-z]', m.group(2)):
        return m.group(1).strip(), m.group(2).strip()
    if re.search(r'[\u4e00-\u9fff]', t) and re.search(r'[A-Za-z]', t):
        return t, ''
    if re.search(r'[\u4e00-\u9fff]', t):
        return t, ''
    return '', t


def find_ref_file(ref_dir: str | None = None) -> str | None:
    root = ref_dir or os.path.join(BASE_DIR, 'Ref')
    if not os.path.isdir(root):
        return None
    candidates = []
    for name in os.listdir(root):
        if not name.lower().endswith('.xlsx') or name.startswith('~$'):
            continue
        if any(h.lower() in name.lower() for h in _REF_NAME_HINTS):
            candidates.append(os.path.join(root, name))
    if not candidates:
        return DEFAULT_REF_PATH if os.path.isfile(DEFAULT_REF_PATH) else None
    candidates.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return candidates[0]


def parse_category_workbook(filepath: str | None = None) -> list[dict]:
    path = filepath or find_ref_file()
    if not path or not os.path.isfile(path):
        raise FileNotFoundError(f'找不到工程標準分類表: {path or DEFAULT_REF_PATH}')

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    records = []
    l1_code = l1_name_zh = l1_name_en = None
    sort_order = 0

    for row in ws.iter_rows(min_row=5, values_only=True):
        if not any(c is not None and _s(c) for c in row[:4]):
            continue
        a, b, c, d, e = (row[i] if i < len(row) else None for i in range(5))
        if a:
            l1_code = _s(a)
            l1_zh, l1_en = _split_bilingual(b)
            l1_name_zh, l1_name_en = l1_zh, l1_en
        if not c:
            continue
        l2_code = _s(c)
        l2_zh, l2_en = _split_bilingual(d)
        sort_order += 1
        records.append({
            'l2_code': l2_code,
            'l1_code': l1_code,
            'l1_name_zh': l1_name_zh,
            'l1_name_en': l1_name_en,
            'l2_name_zh': l2_zh,
            'l2_name_en': l2_en,
            'scope': _s(e) or None,
            'sort_order': sort_order,
            'source_file': os.path.basename(path),
        })

    wb.close()
    if not records:
        raise ValueError(f'未能從 {path} 解析工程分類')
    return records


def _project_text_blob(project: dict) -> str:
    parts = [
        project.get('project_name_zh'),
        project.get('project_name_en'),
        project.get('project_name'),
        project.get('notes'),
        project.get('work_type'),
    ]
    return ' '.join(_s(p) for p in parts if _s(p)).lower()


def suggest_category_l2(project: dict) -> str | None:
    """依項目名稱關鍵字建議二級分類；無匹配則 None"""
    blob = _project_text_blob(project)
    if not blob:
        return None
    wt = _s(project.get('work_type')).lower()
    if wt and wt not in _PLACEHOLDER_WORK_TYPES:
        blob = f'{blob} {wt}'
    for code, keywords in CLASSIFY_RULES:
        for kw in keywords:
            if kw.lower() in blob:
                return code
    return None


def category_display_label(row: dict | None, level: str = 'l2') -> str:
    if not row:
        return '—'
    if level == 'l1':
        zh = _s(row.get('l1_name_zh'))
        en = _s(row.get('l1_name_en'))
        code = _s(row.get('l1_code'))
    else:
        zh = _s(row.get('l2_name_zh'))
        en = _s(row.get('l2_name_en'))
        code = _s(row.get('l2_code'))
    name = zh or en or '—'
    return f'{code} {name}' if code else name


def ref_status(ref_dir: str | None = None) -> dict:
    path = find_ref_file(ref_dir)
    ok = bool(path and os.path.isfile(path))
    return {
        'ok': ok,
        'ref_path': path,
        'ref_file': os.path.basename(path) if path else None,
        'default_path': DEFAULT_REF_PATH,
    }
