"""
sc_contract_ref.py — 分判工程合約編號（QS Ref Excel）

對照 Ref/分判工程合約編號 YYYY.xlsx（2023 起新格式）：
  MS/C月份-序號/年份/負責人  → P1 Sub-Contracts No.（例 MS/C07-1/25/jc）
  外判公司、工程項目、項目編號、負責同事、分判合約合額
  合約會簽表、合作伙伴、定標會議紀錄（2023–2024 另有 Final Account / ISO / Remark）
"""
from __future__ import annotations

import os
import re

from config import BASE_DIR
from sc_ref import company_matches_sc, derive_parent_sc_no

_MS_C_RE = re.compile(r'^MS/C[\d\-A-Za-z/]+$', re.IGNORECASE)
_LEGACY_CONTRACT_RE = re.compile(r'^\d{4}/', re.IGNORECASE)
_REF_NAME_HINTS = ('分判工程合約編號', '合約編號')


def _is_ms_c_contract_no(val) -> bool:
    s = (val or '').strip()
    return bool(s and _MS_C_RE.match(s))


def _is_legacy_contract_no(val) -> bool:
    """2023 前舊格式，例 2018/MP/10/001"""
    s = (val or '').strip()
    if not s or _is_ms_c_contract_no(s):
        return False
    return bool(_LEGACY_CONTRACT_RE.match(s))


def _is_registry_contract_no(val) -> bool:
    return _is_ms_c_contract_no(val) or _is_legacy_contract_no(val)


def _sheet_year(sheet) -> int | None:
    m = re.match(r'(20\d{2})', str(sheet or ''))
    return int(m.group(1)) if m else None


def display_p1_sub_contract_no(val) -> str:
    """
    P1「分判合約編號 / Sub-Contracts No.」顯示值。
    只允許 MS/C… 格式；判項編號（SC-/O-/M-）、報價單號等一律顯示 —。
    """
    s = (val or '').strip()
    if _is_ms_c_contract_no(s):
        return s
    return '—'


def normalize_ms_c_input(val) -> str | None:
    """表單／API 輸入：空白 → None，否則 strip"""
    s = (val or '').strip()
    return s or None


def validate_ms_c_input(val) -> tuple[bool, str]:
    s = normalize_ms_c_input(val)
    if s is None:
        return True, ''
    if _is_ms_c_contract_no(s):
        return True, s
    if _is_legacy_contract_no(s):
        return True, s
    return False, '合約編號格式不正確（例：MS/C07-1/25/jc 或 2018/MP/10/001）'


def extract_person_code_from_ms_c(sub_contract_no) -> str | None:
    """MS/C07-1/25/jc → jc（負責同事縮寫）"""
    s = (sub_contract_no or '').strip()
    if not _is_ms_c_contract_no(s):
        return None
    if '/' not in s:
        return None
    code = s.rsplit('/', 1)[-1].strip()
    return code.lower() if code else None


def _looks_like_project_code(val) -> bool:
    """是否像項目編號（MS_Q… / Q… / MS/Q…），排除負責同事名"""
    s = (val or '').strip()
    if not s:
        return False
    if re.search(r'整\s*$', s):
        return False
    u = s.upper().replace('/', '_').replace('-', '_')
    u = re.sub(r'_+', '_', u)
    if re.search(r'MS_Q\d+_\d+', u):
        return True
    if re.match(r'Q\d+_\d+', u):
        return True
    if re.match(r'MS/Q', s, re.I):
        return True
    return False


def _cell_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _build_registry_row(
    sub_contract_no,
    company,
    works,
    project_code,
    *,
    legacy_contract_no=None,
    person_in_charge=None,
    amount=None,
    countersign=None,
    partner=None,
    tender_minutes=None,
    final_account=None,
    final_account_statement=None,
    iso_flag=None,
    remark=None,
    sheet=None,
    source_file=None,
) -> dict | None:
    sub_contract_no = (sub_contract_no or '').strip()
    project_code = (project_code or '').strip() if project_code else ''
    if not sub_contract_no:
        return None
    if not project_code and not _is_legacy_contract_no(sub_contract_no):
        return None
    return {
        'sub_contract_no': sub_contract_no,
        'legacy_contract_no': _cell_str(legacy_contract_no),
        'company': _cell_str(company),
        'works': _cell_str(works),
        'project_code': project_code or None,
        'project_core': _project_core(project_code) if project_code else None,
        'person_in_charge': _cell_str(person_in_charge),
        'person_code': extract_person_code_from_ms_c(sub_contract_no),
        'amount': _parse_amount(amount),
        'countersign': _cell_str(countersign),
        'partner': _cell_str(partner),
        'tender_minutes': _cell_str(tender_minutes),
        'final_account': _cell_str(final_account),
        'final_account_statement': _cell_str(final_account_statement),
        'iso_flag': _cell_str(iso_flag),
        'remark': _cell_str(remark),
        'sheet': _cell_str(sheet),
        'source_file': _cell_str(source_file),
    }


def _project_core(code) -> str:
    """MS_Q0185_25_kp / MS_Q0185_25 / Q0185_25 → MS_Q0185_25"""
    if not code:
        return ''
    t = str(code).strip().upper().replace('/', '_').replace('-', '_')
    t = re.sub(r'_+', '_', t)
    m = re.match(r'(MS_Q\d+_\d+)', t)
    if m:
        return m.group(1)
    m = re.match(r'Q(\d+_\d+)', t)
    if m:
        return f'MS_Q{m.group(1)}'
    return t


def _project_keys(project: dict) -> set[str]:
    keys = set()
    for raw in (
        project.get('project_code'),
        project.get('mp_contract_code'),
        project.get('quotation_no'),
        project.get('job_no'),
    ):
        core = _project_core(raw)
        if core:
            keys.add(core)
    return keys


def _find_ref_workbook() -> str | None:
    ref_dir = os.path.join(BASE_DIR, 'Ref')
    if not os.path.isdir(ref_dir):
        return None
    candidates = []
    for name in os.listdir(ref_dir):
        if not name.lower().endswith('.xlsx') or name.startswith('~$'):
            continue
        if any(h in name for h in _REF_NAME_HINTS):
            candidates.append(os.path.join(ref_dir, name))
    if not candidates:
        return None

    def year_key(p):
        m = re.search(r'(20\d{2})', os.path.basename(p))
        return int(m.group(1)) if m else 0

    candidates.sort(key=year_key, reverse=True)
    return candidates[0]


def _parse_amount(val) -> float:
    if val is None or val == '':
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(',', '')
    m = re.search(r'[\d.]+', s)
    if not m:
        return 0.0
    try:
        return float(m.group(0))
    except ValueError:
        return 0.0


def _parse_legacy_row(row, sheet) -> dict | None:
    """2023 前舊格式合約編號（A 欄 YYYY/MP/…）"""
    c0 = _cell_str(row[0]) or ''
    if not _is_legacy_contract_no(c0):
        return None
    year = _sheet_year(sheet) or 0
    if year >= 2022:
        return _build_registry_row(
            c0,
            row[1] if len(row) > 1 else None,
            row[2] if len(row) > 2 else None,
            row[3] if len(row) > 3 else None,
            person_in_charge=row[4] if len(row) > 4 else None,
            amount=row[5] if len(row) > 5 else None,
            countersign=row[6] if len(row) > 6 else None,
            partner=row[7] if len(row) > 7 else None,
            tender_minutes=row[8] if len(row) > 8 else None,
            sheet=sheet,
        )
    if year == 2021:
        return _build_registry_row(
            c0,
            row[1] if len(row) > 1 else None,
            row[2] if len(row) > 2 else None,
            row[3] if len(row) > 3 else None,
            person_in_charge=row[5] if len(row) > 5 else None,
            amount=row[6] if len(row) > 6 else None,
            countersign=row[7] if len(row) > 7 else None,
            partner=row[8] if len(row) > 8 else None,
            tender_minutes=row[9] if len(row) > 9 else None,
            sheet=sheet,
        )
    if year == 2020:
        return _build_registry_row(
            c0,
            row[1] if len(row) > 1 else None,
            row[2] if len(row) > 2 else None,
            None,
            person_in_charge=row[3] if len(row) > 3 else None,
            countersign=row[4] if len(row) > 4 else None,
            partner=row[5] if len(row) > 5 else None,
            tender_minutes=row[6] if len(row) > 6 else None,
            sheet=sheet,
        )
    # 2018–2019：僅 合約編號／外判公司／工程項目／負責同事（D 欄無項目編號）
    person = row[3] if len(row) > 3 else None
    return _build_registry_row(
        c0,
        row[1] if len(row) > 1 else None,
        row[2] if len(row) > 2 else None,
        None,
        person_in_charge=person,
        sheet=sheet,
    )


def _parse_sheet_row(row, sheet=None) -> dict | None:
    """支援 MS/C（2023+）、過渡（2023–24）、舊編號（2022 前）"""
    if not row:
        return None
    c0 = _cell_str(row[0]) or ''
    c1 = _cell_str(row[1]) if len(row) > 1 else ''
    if _is_ms_c_contract_no(c0):
        parsed = _build_registry_row(
            c0,
            row[1] if len(row) > 1 else None,
            row[2] if len(row) > 2 else None,
            row[3] if len(row) > 3 else None,
            person_in_charge=row[4] if len(row) > 4 else None,
            amount=row[5] if len(row) > 5 else None,
            countersign=row[6] if len(row) > 6 else None,
            partner=row[7] if len(row) > 7 else None,
            tender_minutes=row[8] if len(row) > 8 else None,
            final_account=row[9] if len(row) > 9 else None,
            final_account_statement=row[10] if len(row) > 10 else None,
            iso_flag=row[11] if len(row) > 11 else None,
            remark=row[12] if len(row) > 12 else None,
            sheet=sheet,
        )
        if parsed:
            return parsed
    if _is_ms_c_contract_no(c1):
        parsed = _build_registry_row(
            c1,
            row[2] if len(row) > 2 else None,
            row[3] if len(row) > 3 else None,
            row[4] if len(row) > 4 else None,
            legacy_contract_no=c0 or None,
            person_in_charge=row[5] if len(row) > 5 else None,
            amount=row[6] if len(row) > 6 else None,
            countersign=row[7] if len(row) > 7 else None,
            partner=row[8] if len(row) > 8 else None,
            tender_minutes=row[9] if len(row) > 9 else None,
            final_account=row[10] if len(row) > 10 else None,
            final_account_statement=row[11] if len(row) > 11 else None,
            iso_flag=row[12] if len(row) > 12 else None,
            remark=row[13] if len(row) > 13 else None,
            sheet=sheet,
        )
        if parsed:
            return parsed
    return _parse_legacy_row(row, sheet)


def _load_rows_from_workbook(path: str) -> list[dict]:
    import openpyxl
    rows = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            for row in ws.iter_rows(min_row=2, values_only=True):
                parsed = _parse_sheet_row(row, sheet=sheet)
                if parsed:
                    if not parsed.get('sheet'):
                        parsed['sheet'] = sheet
                    rows.append(parsed)
    finally:
        wb.close()
    return rows


_cache: dict = {'path': None, 'mtime': 0.0, 'rows': []}


def _load_ref_rows() -> list[dict]:
    try:
        import database as db
        db_rows = db.list_sc_contract_registry_rows()
        if db_rows:
            return db_rows
    except Exception:
        pass
    path = _find_ref_workbook()
    if not path or not os.path.isfile(path):
        _cache.update({'path': None, 'mtime': 0.0, 'rows': []})
        return []
    mtime = os.path.getmtime(path)
    if _cache['path'] == path and _cache['mtime'] == mtime and _cache['rows']:
        return _cache['rows']
    rows = _load_rows_from_workbook(path)
    _cache.update({'path': path, 'mtime': mtime, 'rows': rows})
    return rows


def invalidate_cache():
    _cache.update({'path': None, 'mtime': 0.0, 'rows': []})


def ref_status() -> dict:
    path = _find_ref_workbook()
    rows = _load_ref_rows()
    try:
        import database as db
        db_status = db.get_sc_contract_registry_status()
    except Exception:
        db_status = {'row_count': 0, 'last_import': None}
    return {
        'path': path,
        'row_count': len(rows),
        'loaded': bool(rows),
        'db_row_count': db_status.get('row_count', 0),
        'last_import': db_status.get('last_import'),
        'source': 'database' if db_status.get('row_count') else 'excel',
    }


def classify_contract_align(sc: dict, resolved: str) -> str:
    """aligned | missing | na"""
    if resolved and resolved != '—':
        return 'aligned'
    sn = (sc.get('sc_no') or '').strip().upper()
    if sn.startswith('O-') or sn.startswith('M-'):
        return 'na'
    return 'missing'


def project_orphan_refs(project: dict, project_subcontractors: list | None) -> list[dict]:
    """Excel/DB 有 MS/C，但本項目判項未配對到的記錄"""
    keys = _project_keys(project)
    if not keys:
        return []
    used = set()
    for sc in project_subcontractors or []:
        if sc.get('is_excluded'):
            continue
        no = resolve_sub_contract_no(project, sc, project_subcontractors)
        if no and no != '—':
            used.add(no)
    orphans = []
    for row in _load_ref_rows():
        if row['project_core'] not in keys:
            continue
        if row['sub_contract_no'] in used:
            continue
        orphans.append({
            'sub_contract_no': row['sub_contract_no'],
            'company': row.get('company'),
            'works': (row.get('works') or '')[:80],
            'project_code': row.get('project_code'),
            'amount': row.get('amount'),
        })
    return orphans


def _norm_text(val) -> str:
    return re.sub(r'\s+', '', str(val or '').strip().lower())


def _sc_works_text(sc: dict) -> str:
    parts = [
        sc.get('description'),
        sc.get('trade_label'),
        sc.get('service_description'),
        sc.get('scope'),
    ]
    return ' '.join(str(p).strip() for p in parts if p)


def _works_score(sc: dict, row: dict) -> int:
    """工程項目相似度（同項目編號、不同外判時輔助配對）"""
    ref_works = _norm_text(row.get('works'))
    if not ref_works:
        return 0
    sc_text = _norm_text(_sc_works_text(sc))
    if not sc_text:
        return 0
    score = 0
    if ref_works in sc_text or sc_text in ref_works:
        score += 12
    for token in re.findall(r'[\u4e00-\u9fff]{2,}|[a-zA-Z]{3,}', row.get('works') or ''):
        t = _norm_text(token)
        if len(t) >= 2 and t in sc_text:
            score += 2
    return score


def _match_rows(project: dict, sc: dict, rows: list[dict]) -> list[dict]:
    keys = _project_keys(project)
    if not keys:
        return []
    candidates = [row for row in rows if row['project_core'] in keys]
    if not candidates:
        return []

    # 同項目編號可能有多個外判 → 先以公司名篩選
    by_company = [row for row in candidates if company_matches_sc(sc, row['company'])]
    if by_company:
        return by_company

    # 公司名對不上 DB 時，以工程項目文字輔助（仍需同項目編號）
    scored = [(row, _works_score(sc, row)) for row in candidates]
    scored = [(row, score) for row, score in scored if score > 0]
    if not scored:
        return []
    best = max(score for _, score in scored)
    return [row for row, score in scored if score >= best - 1]


def _pick_best_row(sc: dict, matched: list[dict]) -> dict | None:
    if not matched:
        return None
    if len(matched) == 1:
        return matched[0]

    amt = float(sc.get('contract_amount') or sc.get('contract_sum') or 0)

    def rank(row):
        works = _works_score(sc, row)
        amt_diff = abs(row['amount'] - amt) if amt > 0 else float('inf')
        return (-works, amt_diff)

    return min(matched, key=rank)


def resolve_sub_contract_no(
    project: dict,
    sc: dict,
    project_subcontractors: list | None = None,
    *,
    _depth: int = 0,
) -> str:
    """
    P1 分判合約編號（Sub-Contracts No.）
    優先：DB 手動欄位 → MS/C 格式 quotation_no → Ref Excel → 父判項繼承 → —
    """
    if _depth > 3:
        return '—'

    manual = (sc.get('sub_contract_no') or '').strip()
    if manual:
        return manual

    qn = (sc.get('quotation_no') or '').strip()
    if _is_ms_c_contract_no(qn):
        return qn

    rows = _load_ref_rows()
    matched = _match_rows(project, sc, rows)
    best = _pick_best_row(sc, matched)
    if best:
        return best['sub_contract_no']

    if _depth == 0 and project_subcontractors:
        parent_no = derive_parent_sc_no(sc.get('sc_no'))
        cur_no = (sc.get('sc_no') or '').strip()
        if parent_no and parent_no != cur_no:
            for psc in project_subcontractors:
                if (psc.get('sc_no') or '').strip() == parent_no:
                    inherited = resolve_sub_contract_no(
                        project, psc, project_subcontractors, _depth=_depth + 1,
                    )
                    if inherited and inherited != '—':
                        return inherited
                    break

    return '—'
